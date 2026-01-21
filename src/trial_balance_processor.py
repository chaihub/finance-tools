"""Trial Balance processor for Manufacturing Accounting workbook."""

from typing import Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class TrialBalanceProcessor:
    """Processes and populates the Trial Balance sheet from source data."""

    def __init__(self, file_path: str):
        """
        Initialize the TrialBalanceProcessor.

        Args:
            file_path (str): Path to the Excel file.
        """
        self.file_path = file_path
        self.workbook = load_workbook(file_path)

    def populate_trial_balance(self) -> None:
        """
        Populate the Trial Balance sheet with Account Code, Account Name, Debit, and Credit.

        Workflow:
        1. Load Chart_of_Accounts and populate Account Code and Account Name
        2. Load General_Ledger and sum Debit/Credit amounts by Account Code
        3. Write results to Trial_Balance sheet
        """
        # Get source sheets
        chart_of_accounts = self.workbook["Chart_of_Accounts"]
        general_ledger = self.workbook["General_Ledger"]
        trial_balance = self.workbook["Trial_Balance"]

        # Step 1: Extract Chart of Accounts data
        accounts_map = self._load_chart_of_accounts(chart_of_accounts)

        # Step 2: Aggregate debits and credits from General Ledger
        ledger_summary = self._aggregate_ledger(general_ledger)

        # Step 3: Populate Trial Balance sheet
        self._write_trial_balance(trial_balance, accounts_map, ledger_summary)

        # Save the workbook
        self.workbook.save(self.file_path)

    def _load_chart_of_accounts(
        self, sheet: Worksheet
    ) -> Dict[str, str]:
        """
        Load the Chart of Accounts mapping.

        Args:
            sheet (Worksheet): The Chart_of_Accounts sheet.

        Returns:
            Dict[str, str]: Mapping of Account Code to Account Name.
        """
        accounts_map = {}
        # Skip header row (row 1)
        for row in sheet.iter_rows(min_row=2, values_only=False):
            account_code = row[0].value
            account_name = row[1].value
            if account_code:
                accounts_map[account_code] = account_name
        return accounts_map

    def _aggregate_ledger(self, sheet: Worksheet) -> Dict[str, Tuple[float, float]]:
        """
        Aggregate debit and credit amounts from General Ledger by Account Code.

        Args:
            sheet (Worksheet): The General_Ledger sheet.

        Returns:
            Dict[str, Tuple[float, float]]: Mapping of Account Code to (Debit, Credit) totals.
        """
        ledger_summary = {}
        # Skip header row (row 1)
        for row in sheet.iter_rows(min_row=2, values_only=False):
            account_code = row[2].value  # Column C: Account Code
            debit = row[3].value or 0  # Column D: Debit
            credit = row[4].value or 0  # Column E: Credit

            if account_code:
                if account_code not in ledger_summary:
                    ledger_summary[account_code] = (0, 0)
                current_debit, current_credit = ledger_summary[account_code]
                ledger_summary[account_code] = (
                    current_debit + debit,
                    current_credit + credit,
                )
        return ledger_summary

    def _write_trial_balance(
        self,
        sheet: Worksheet,
        accounts_map: Dict[str, str],
        ledger_summary: Dict[str, Tuple[float, float]],
    ) -> None:
        """
        Write aggregated data to the Trial Balance sheet.

        Args:
            sheet (Worksheet): The Trial_Balance sheet.
            accounts_map (Dict[str, str]): Account Code to Account Name mapping.
            ledger_summary (Dict[str, Tuple[float, float]]): Account Code to (Debit, Credit).
        """
        row_num = 2  # Start from row 2 (row 1 is header)
        for account_code in sorted(accounts_map.keys()):
            account_name = accounts_map[account_code]
            debit, credit = ledger_summary.get(account_code, (0, 0))

            sheet.cell(row=row_num, column=1, value=account_code)  # Account Code
            sheet.cell(row=row_num, column=2, value=account_name)  # Account Name
            sheet.cell(row=row_num, column=3, value=debit)  # Debit
            sheet.cell(row=row_num, column=4, value=credit)  # Credit

            row_num += 1

    def populate_income_statement(self) -> None:
        """
        Populate the Income Statement sheet from Trial Balance data.

        Calculates:
        - Sales Revenue: from account 4000 (Credit - Debit)
        - Cost of Goods Sold: from account 5000 (Debit - Credit)
        - Gross Profit: Sales Revenue - COGS
        - Salaries Expense: from account 5100 (Debit - Credit)
        - Net Profit: Gross Profit - Salaries Expense
        """
        trial_balance = self.workbook["Trial_Balance"]
        income_statement = self.workbook["Income_Statement"]

        # Build Trial Balance data map
        tb_data = {}
        for row in trial_balance.iter_rows(min_row=2, values_only=False):
            account_code = row[0].value
            debit = row[2].value or 0
            credit = row[3].value or 0
            if account_code:
                tb_data[account_code] = (debit, credit)

        # Calculate amounts
        sales_revenue = tb_data.get(4000, (0, 0))[1] - tb_data.get(4000, (0, 0))[0]  # Credit - Debit
        cogs = tb_data.get(5000, (0, 0))[0] - tb_data.get(5000, (0, 0))[1]  # Debit - Credit
        salaries = tb_data.get(5100, (0, 0))[0] - tb_data.get(5100, (0, 0))[1]  # Debit - Credit
        gross_profit = sales_revenue - cogs
        net_profit = gross_profit - salaries

        # Write to Income Statement (rows 2-6 correspond to the line items)
        income_statement.cell(row=2, column=2, value=sales_revenue)
        income_statement.cell(row=3, column=2, value=cogs)
        income_statement.cell(row=4, column=2, value=gross_profit)
        income_statement.cell(row=5, column=2, value=salaries)
        income_statement.cell(row=6, column=2, value=net_profit)

        # Save the workbook
        self.workbook.save(self.file_path)

    def populate_balance_sheet(self) -> None:
        """
        Populate the Balance_Sheet sheet from Trial Balance data.

        Balance Sheet amounts are net of debits and credits.
        For all accounts: Amount = Credit - Debit
        """
        trial_balance = self.workbook["Trial_Balance"]
        balance_sheet = self.workbook["Balance_Sheet"]

        # Build Trial Balance data map
        tb_data = {}
        for row in trial_balance.iter_rows(min_row=2, values_only=False):
            account_code = row[0].value
            account_name = row[1].value
            debit = row[2].value or 0
            credit = row[3].value or 0
            if account_code:
                tb_data[account_code] = (debit, credit)

        # Asset accounts: 1000 (Cash), 1100 (A/R), 1200 (Inventory), 1500 (Plant & Machinery)
        assets_map = {
            "Cash at Bank": 1000,
            "Inventory": 1200,
            "Accounts Receivable": 1100,
        }

        # Liability & Equity accounts: 3000 (Capital), 2000 (A/P), 3100 (Retained Earnings)
        liabilities_equity_map = {
            "Equity Share Capital": 3000,
            "Accounts Payable": 2000,
            "Retained Earnings": 3100,
        }

        # Write Assets (rows 2-4, column 2)
        asset_rows = [2, 3, 4]
        for row_idx, (asset_name, account_code) in enumerate(assets_map.items()):
            debit, credit = tb_data.get(account_code, (0, 0))
            amount = credit - debit
            balance_sheet.cell(row=asset_rows[row_idx], column=2, value=amount)

        # Write Liabilities & Equity (rows 2-4, column 4)
        le_rows = [2, 3, 4]
        for row_idx, (le_name, account_code) in enumerate(liabilities_equity_map.items()):
            debit, credit = tb_data.get(account_code, (0, 0))
            amount = credit - debit
            balance_sheet.cell(row=le_rows[row_idx], column=4, value=amount)

        # Save the workbook
        self.workbook.save(self.file_path)
