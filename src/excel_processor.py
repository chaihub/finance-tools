# src/excel_processor.py

import logging
from typing import Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelProcessor:
    """
    Main class to orchestrate the Excel processing workflow.
    """
    
    def __init__(self, file_path: str):
        """
        Initialize the ExcelProcessor with the file path.
        
        Args:
            file_path (str): Path to the Excel file.
        """
        self.file_path = file_path
        self.excel_reader = ExcelReader(file_path)
        self.calculator = Calculator()
        self.excel_writer = ExcelWriter(file_path)
        self.error_handler = ErrorHandler()
        
    def process(self, input_sheet: str, output_sheet: str, calculations: Dict[str, Any]) -> None:
        """
        Orchestrate the Excel processing workflow.
        
        Args:
            input_sheet (str): Name of the input sheet.
            output_sheet (str): Name of the output sheet.
            calculations (Dict[str, Any]): Dictionary of calculations to perform.
        """
        try:
            # Read data from the input sheet
            data = self.excel_reader.read_sheet(input_sheet)
            
            # Perform calculations
            results = self.calculator.calculate(data, calculations)
            
            # Write results to the output sheet
            self.excel_writer.write_sheet(output_sheet, results)
            
            logging.info("Processing completed successfully.")
        except Exception as e:
            self.error_handler.handle_error(e)
            raise


class ExcelReader:
    """
    Class to read data from Excel files.
    """
    
    def __init__(self, file_path: str):
        """
        Initialize the ExcelReader with the file path.
        
        Args:
            file_path (str): Path to the Excel file.
        """
        self.file_path = file_path
        
    def read_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """
        Read data from a specified sheet.
        
        Args:
            sheet_name (str): Name of the sheet to read.
        
        Returns:
            Dict[str, Any]: Data read from the sheet.
        """
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook[sheet_name]
            data = self._parse_sheet(sheet)
            return data
        except Exception as e:
            ErrorHandler().handle_error(e)
            raise
    
    def _parse_sheet(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        Parse the sheet data into a dictionary.
        
        Args:
            sheet (Worksheet): The worksheet to parse.
        
        Returns:
            Dict[str, Any]: Parsed data.
        """
        data = {}
        for row in sheet.iter_rows(values_only=True):
            if row and row[0]:
                data[row[0]] = row[1:]
        return data


class Calculator:
    """
    Class to perform calculations on the data.
    """
    
    def calculate(self, data: Dict[str, Any], calculations: Dict[str, Any]) -> Dict[str, Any]:
        """
        Perform calculations on the data.
        
        Args:
            data (Dict[str, Any]): Input data.
            calculations (Dict[str, Any]): Dictionary of calculations to perform.
        
        Returns:
            Dict[str, Any]: Results of the calculations.
        """
        results = {}
        for key, value in calculations.items():
            if key in data:
                results[key] = self._perform_calculation(data[key], value)
        return results
    
    def _perform_calculation(self, input_data: Any, calculation: Any) -> Any:
        """
        Perform a single calculation.
        
        Args:
            input_data (Any): Input data for the calculation.
            calculation (Any): Calculation to perform.
        
        Returns:
            Any: Result of the calculation.
        """
        # Example calculation: sum of all values
        if calculation == "sum":
            return sum(input_data)
        elif calculation == "average":
            return sum(input_data) / len(input_data)
        else:
            raise ValueError(f"Unsupported calculation: {calculation}")


class ExcelWriter:
    """
    Class to write data back to Excel files.
    """
    
    def __init__(self, file_path: str):
        """
        Initialize the ExcelWriter with the file path.
        
        Args:
            file_path (str): Path to the Excel file.
        """
        self.file_path = file_path
        
    def write_sheet(self, sheet_name: str, data: Dict[str, Any]) -> None:
        """
        Write data to a specified sheet.
        
        Args:
            sheet_name (str): Name of the sheet to write to.
            data (Dict[str, Any]): Data to write.
        """
        try:
            workbook = load_workbook(self.file_path)
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.create_sheet(sheet_name)
            
            self._write_data(sheet, data)
            workbook.save(self.file_path)
        except Exception as e:
            ErrorHandler().handle_error(e)
            raise
    
    def _write_data(self, sheet: Worksheet, data: Dict[str, Any]) -> None:
        """
        Write data to the sheet.
        
        Args:
            sheet (Worksheet): The worksheet to write to.
            data (Dict[str, Any]): Data to write.
        """
        for row_idx, (key, values) in enumerate(data.items(), start=1):
            sheet.cell(row=row_idx, column=1, value=key)
            if isinstance(values, (list, tuple)):
                for col_idx, value in enumerate(values, start=2):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            else:
                sheet.cell(row=row_idx, column=2, value=values)


class ErrorHandler:
    """
    Class to handle errors and logging.
    """
    
    def __init__(self):
        """
        Initialize the ErrorHandler.
        """
        logging.basicConfig(level=logging.INFO)
        
    def handle_error(self, error: Exception) -> None:
        """
        Handle errors and log them.
        
        Args:
            error (Exception): The error to handle.
        """
        logging.error(f"An error occurred: {error}")
        raise error