"""Tests for the TrialBalanceProcessor class."""

import os
import shutil
import pytest
from src.trial_balance_processor import TrialBalanceProcessor


@pytest.fixture
def test_workbook():
    """Copy the existing Manufacturing_Accounting_Simple.xlsx and return the copy path."""
    source_file = "data/Manufacturing_Accounting_Simple.xlsx"
    test_file = "data/test_trial_balance_output.xlsx"

    # Copy the existing file
    shutil.copy(source_file, test_file)
    yield test_file

    # Note: Copy is not deleted - retained for inspection


def test_populate_trial_balance(test_workbook):
    """Test that Trial_Balance sheet is correctly populated with real manufacturing data."""
    processor = TrialBalanceProcessor(test_workbook)
    processor.populate_trial_balance()

    # Reload the workbook to verify results
    from openpyxl import load_workbook

    wb = load_workbook(test_workbook)
    trial_sheet = wb["Trial_Balance"]

    # Verify data was written
    assert trial_sheet.max_row > 1, "Trial Balance sheet should have data rows"

    # Check that all accounts from Chart_of_Accounts are present
    account_codes = []
    for row in trial_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            account_codes.append(row[0])

    expected_codes = [1000, 1100, 1200, 1500, 2000, 3000, 3100, 4000, 5000, 5100, 5200]
    assert sorted(account_codes) == sorted(expected_codes), "All accounts should be populated"

    # Build a map of account codes to data
    trial_data = {}
    for row in trial_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            trial_data[row[0]] = {"name": row[1], "debit": row[2] or 0, "credit": row[3] or 0}

    # Verify account names were populated correctly
    assert trial_data[1000]["name"] == "Cash at Bank"
    assert trial_data[1100]["name"] == "Accounts Receivable"
    assert trial_data[2000]["name"] == "Accounts Payable"
    assert trial_data[3000]["name"] == "Equity Share Capital"

    # Verify that debits and credits are populated (exact values depend on ledger)
    # Just verify that at least some accounts have non-zero values
    total_debits = sum(row["debit"] for row in trial_data.values())
    total_credits = sum(row["credit"] for row in trial_data.values())
    assert total_debits > 0, "Should have some debits"
    assert total_credits > 0, "Should have some credits"


@pytest.fixture
def test_workbook_for_income():
    """Copy the existing Manufacturing_Accounting_Simple.xlsx for income statement test."""
    source_file = "data/Manufacturing_Accounting_Simple.xlsx"
    test_file = "data/test_income_statement_output.xlsx"

    # Copy the existing file
    shutil.copy(source_file, test_file)
    yield test_file

    # Note: Copy is not deleted - retained for inspection


def test_populate_income_statement(test_workbook_for_income):
    """Test that Income_Statement sheet is correctly populated from Trial Balance."""
    processor = TrialBalanceProcessor(test_workbook_for_income)

    # First populate Trial Balance
    processor.populate_trial_balance()

    # Then populate Income Statement
    processor.populate_income_statement()

    # Reload the workbook to verify results
    from openpyxl import load_workbook

    wb = load_workbook(test_workbook_for_income)
    tb_sheet = wb["Trial_Balance"]
    is_sheet = wb["Income_Statement"]

    # Build Trial Balance data for verification
    tb_data = {}
    for row in tb_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            tb_data[row[0]] = {"debit": row[2] or 0, "credit": row[3] or 0}

    # Read Income Statement values
    is_data = {}
    descriptions = ["Sales Revenue", "Cost of Goods Sold", "Gross Profit", "Salaries Expense", "Net Profit"]
    for i, desc in enumerate(descriptions, start=2):
        is_data[desc] = is_sheet.cell(row=i, column=2).value

    # Verify Sales Revenue (account 4000: credit - debit)
    expected_sales_revenue = tb_data[4000]["credit"] - tb_data[4000]["debit"]
    assert is_data["Sales Revenue"] == expected_sales_revenue, f"Sales Revenue should be {expected_sales_revenue}"

    # Verify COGS (account 5000: debit - credit)
    expected_cogs = tb_data[5000]["debit"] - tb_data[5000]["credit"]
    assert is_data["Cost of Goods Sold"] == expected_cogs, f"COGS should be {expected_cogs}"

    # Verify Salaries Expense (account 5100: debit - credit)
    expected_salaries = tb_data[5100]["debit"] - tb_data[5100]["credit"]
    assert is_data["Salaries Expense"] == expected_salaries, f"Salaries should be {expected_salaries}"

    # Verify Gross Profit
    expected_gross_profit = expected_sales_revenue - expected_cogs
    assert is_data["Gross Profit"] == expected_gross_profit, f"Gross Profit should be {expected_gross_profit}"

    # Verify Net Profit
    expected_net_profit = expected_gross_profit - expected_salaries
    assert is_data["Net Profit"] == expected_net_profit, f"Net Profit should be {expected_net_profit}"

    # Verify all values are populated (not None)
    for desc, value in is_data.items():
        assert value is not None, f"{desc} should be populated"


@pytest.fixture
def test_workbook_for_balance_sheet():
    """Copy the existing Manufacturing_Accounting_Simple.xlsx for balance sheet test."""
    source_file = "data/Manufacturing_Accounting_Simple.xlsx"
    test_file = "data/test_balance_sheet_output.xlsx"

    # Copy the existing file
    shutil.copy(source_file, test_file)
    yield test_file

    # Note: Copy is not deleted - retained for inspection


def test_populate_balance_sheet(test_workbook_for_balance_sheet):
    """Test that Balance_Sheet sheet is correctly populated from Trial Balance."""
    processor = TrialBalanceProcessor(test_workbook_for_balance_sheet)

    # First populate Trial Balance
    processor.populate_trial_balance()

    # Then populate Balance Sheet
    processor.populate_balance_sheet()

    # Reload the workbook to verify results
    from openpyxl import load_workbook

    wb = load_workbook(test_workbook_for_balance_sheet)
    tb_sheet = wb["Trial_Balance"]
    bs_sheet = wb["Balance_Sheet"]

    # Build Trial Balance data for verification
    tb_data = {}
    for row in tb_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            tb_data[row[0]] = {"debit": row[2] or 0, "credit": row[3] or 0}

    # Read Balance Sheet values
    # Assets are in rows 2-4, column 2
    # Liabilities & Equity are in rows 2-4, column 4
    assets = {
        "Cash at Bank": bs_sheet.cell(row=2, column=2).value,
        "Inventory": bs_sheet.cell(row=3, column=2).value,
        "Accounts Receivable": bs_sheet.cell(row=4, column=2).value,
    }

    liabilities_equity = {
        "Equity Share Capital": bs_sheet.cell(row=2, column=4).value,
        "Accounts Payable": bs_sheet.cell(row=3, column=4).value,
        "Retained Earnings": bs_sheet.cell(row=4, column=4).value,
    }

    # Verify Asset amounts (Credit - Debit from Trial Balance)
    expected_cash = tb_data[1000]["credit"] - tb_data[1000]["debit"]
    expected_inventory = tb_data[1200]["credit"] - tb_data[1200]["debit"]
    expected_ar = tb_data[1100]["credit"] - tb_data[1100]["debit"]

    assert assets["Cash at Bank"] == expected_cash, f"Cash at Bank should be {expected_cash}"
    assert assets["Inventory"] == expected_inventory, f"Inventory should be {expected_inventory}"
    assert assets["Accounts Receivable"] == expected_ar, f"A/R should be {expected_ar}"

    # Verify Liability & Equity amounts (Credit - Debit from Trial Balance)
    expected_equity = tb_data[3000]["credit"] - tb_data[3000]["debit"]
    expected_ap = tb_data[2000]["credit"] - tb_data[2000]["debit"]
    expected_retained = tb_data[3100]["credit"] - tb_data[3100]["debit"]

    assert (
        liabilities_equity["Equity Share Capital"] == expected_equity
    ), f"Equity Share Capital should be {expected_equity}"
    assert (
        liabilities_equity["Accounts Payable"] == expected_ap
    ), f"Accounts Payable should be {expected_ap}"
    assert (
        liabilities_equity["Retained Earnings"] == expected_retained
    ), f"Retained Earnings should be {expected_retained}"

    # Verify all values are populated (not None)
    for asset_name, amount in assets.items():
        assert amount is not None, f"{asset_name} should be populated"
    for le_name, amount in liabilities_equity.items():
        assert amount is not None, f"{le_name} should be populated"
