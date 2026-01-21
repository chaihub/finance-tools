"""Tests for the ExcelProcessor class."""

import os
import pytest
from openpyxl import Workbook, load_workbook
from src.excel_processor import ExcelProcessor


@pytest.fixture
def excel_file():
    """
    Set up the test environment.
    """
    test_file = "test_data.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Input"
    sheet["A1"] = "Value1"
    sheet["B1"] = 10
    sheet["A2"] = "Value2"
    sheet["B2"] = 20
    workbook.save(test_file)
    yield test_file

    # Clean up the test environment
    if os.path.exists(test_file):
        os.remove(test_file)


def test_process(excel_file):  # pylint: disable=redefined-outer-name
    """
    Test the process method of ExcelProcessor.
    """
    test_file = excel_file
    processor = ExcelProcessor(test_file)
    calculations = {
        "Value1": "sum",
        "Value2": "average"
    }
    processor.process("Input", "Output", calculations)

    # Verify the output sheet was created
    workbook = load_workbook(test_file)
    assert "Output" in workbook.sheetnames

    # Verify the calculations
    output_sheet = workbook["Output"]
    assert output_sheet["B1"].value == 10  # Sum of Value1
    assert output_sheet["B2"].value == 20  # Average of Value2
